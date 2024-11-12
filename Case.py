import openpyxl
import pandas as pd
import google.generativeai as genai
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

genai.configure(api_key="AIzaSyD1q-lK71QDHpDnbEe_1zVVa9B45Tjp4-E")

model = genai.GenerativeModel('gemini-1.5-pro')

# Carregar a planilha original com openpyxl para preservar a formatação
input_path = "C:/Users/ander/Downloads/SUZANO_Exemplos_Indicadores.xlsx"
wb = openpyxl.load_workbook(input_path)
ws = wb.active  # Ou escolha a aba específica, ex: wb['Sheet1']

# Carregar os dados com pandas para manipulação
df = pd.read_excel(input_path)


# Função para classificar o indicador usando a API
def classificar_indicador(indicador):
    try:
        response = model.generate_content(f"Classifique o indicador '{indicador}' como 'Água', 'Energia' ou 'Outros'.")

        if hasattr(response, 'text'):
            classificacao = response.text.strip()  # Remove espaços extras
        else:
            print(f"Erro: Resposta inválida da API para o indicador {indicador}")
            return "Erro"

        if classificacao in ["Água", "Energia", "Outros"]:
            return classificacao
        else:
            print(f"Classificação inválida para o indicador {indicador}: {classificacao}")
            return "Outros"
    except Exception as e:
        print(f"Erro na API para o indicador {indicador}: {e}")
        return "Erro"


# Adicionar a nova coluna "Classificação" à planilha
# Encontrar o número da última coluna
last_column = len(ws[1]) + 1  # Número da última coluna (assumindo que a primeira linha tem os cabeçalhos)

# Adicionar o cabeçalho "Classificação" na última coluna (linha 1)
ws.cell(row=1, column=last_column, value="Classificação")

# Formatação do cabeçalho (linha 1)
header_cell = ws.cell(row=1, column=last_column)
header_cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinha o texto no centro
header_cell.font = Font(color="FFFFFF", bold=True)  # Fonte branca e negrito
header_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Fundo preto
header_cell.border = Border(left=Side(border_style="thin", color="CCCCCC"),
                            right=Side(border_style="thin", color="CCCCCC"),
                            top=Side(border_style="thin", color="CCCCCC"),
                            bottom=Side(border_style="thin", color="CCCCCC"))  # Bordas cinzas

# Aplicar a classificação em cada linha do DataFrame
df["Classificação"] = df["Nome do Indicador"].apply(classificar_indicador)

# Preencher a coluna "Classificação" na planilha com openpyxl e aplicar formatação para os dados (linha 2 em diante)
for row_num, value in enumerate(df["Classificação"], start=2):  # Começando da linha 2
    cell = ws.cell(row=row_num, column=last_column, value=value)

    # Formatação para as células de dados
    cell.alignment = Alignment(horizontal="center", vertical="center")  # Centraliza o conteúdo
    cell.font = Font(color="000000")  # Fonte preta
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Fundo branco
    cell.border = Border(left=Side(border_style="thin", color="CCCCCC"),
                         right=Side(border_style="thin", color="CCCCCC"),
                         top=Side(border_style="thin", color="CCCCCC"),
                         bottom=Side(border_style="thin", color="CCCCCC"))  # Bordas #CCCCCC

# Ajustar a largura da coluna "Classificação"
ws.column_dimensions[openpyxl.utils.get_column_letter(last_column)].width = 35.11

# Salvar a planilha original com a nova coluna "Classificação"
wb.save(input_path)

print(f"Planilha original com a nova coluna 'Classificação' salva em: {input_path}")
